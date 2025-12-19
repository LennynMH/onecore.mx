"""
Excel Exporter - Utilidad para exportar datos a Excel.

Refactorización con Auto (Claude/ChatGPT) - PARTE 3.1

¿Qué hace este módulo?
Proporciona una clase dedicada para exportar datos a archivos Excel con formato,
estilos y auto-ajuste de columnas. Esta refactorización extrae la lógica de
exportación de HistoryUseCases para mejorar la reutilización.

¿Qué clases contiene?
- ExcelExporter: Clase para exportar datos estructurados a Excel
"""

import logging
from typing import List, Dict, Any, Optional
from io import BytesIO
from datetime import datetime

logger = logging.getLogger(__name__)


class ExcelExporter:
    """
    Exportador de datos a archivos Excel.
    
    ¿Qué hace la clase?
    Proporciona métodos para crear archivos Excel con formato profesional,
    incluyendo estilos de encabezados, auto-ajuste de columnas, y manejo
    de diferentes tipos de datos.
    
    ¿Qué métodos tiene?
    - export_events: Exporta eventos a Excel
    - _create_workbook: Crea un workbook con estilos
    - _add_headers: Agrega encabezados con formato
    - _add_data_rows: Agrega filas de datos
    - _auto_adjust_columns: Ajusta ancho de columnas automáticamente
    """
    
    @staticmethod
    def export_events(
        events: List[Dict[str, Any]],
        include_document_details: bool = True,
        sheet_name: str = "Historial de Eventos"
    ) -> BytesIO:
        """
        Exporta eventos a un archivo Excel.
        
        ¿Qué hace la función?
        Crea un archivo Excel con los eventos proporcionados, incluyendo
        encabezados formateados, datos estructurados, y auto-ajuste de columnas.
        
        ¿Qué parámetros recibe y de qué tipo?
        - events (List[Dict[str, Any]]): Lista de eventos a exportar
        - include_document_details (bool): Si True, incluye detalles del documento (default: True)
        - sheet_name (str): Nombre de la hoja de cálculo (default: "Historial de Eventos")
        
        ¿Qué dato regresa y de qué tipo?
        - BytesIO: Buffer con el contenido del archivo Excel
        
        Raises:
            Exception: Si openpyxl no está instalado o si ocurre un error
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            logger.error("openpyxl not installed. Install it with: pip install openpyxl")
            raise Exception("Excel export requires openpyxl library. Please install it.")
        
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # Definir encabezados
        headers = ["ID", "Tipo de Evento", "Descripción", "Fecha y Hora"]
        if include_document_details:
            headers.extend(["ID Documento", "Nombre Archivo", "Clasificación"])
        headers.append("ID Usuario")
        
        # Agregar encabezados con formato
        ExcelExporter._add_headers(ws, headers)
        
        # Agregar filas de datos
        ExcelExporter._add_event_rows(ws, events, include_document_details, start_row=2)
        
        # Auto-ajustar columnas
        ExcelExporter._auto_adjust_columns(ws)
        
        # Guardar en BytesIO
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        logger.info(f"Excel export created: {len(events)} events")
        return excel_buffer
    
    @staticmethod
    def _add_headers(ws, headers: List[str]) -> None:
        """
        Agrega encabezados con formato a la hoja de cálculo.
        
        ¿Qué hace la función?
        Crea la fila de encabezados con estilos profesionales (fondo azul,
        texto blanco en negrita, centrado).
        
        ¿Qué parámetros recibe y de qué tipo?
        - ws: Worksheet de openpyxl
        - headers (List[str]): Lista de nombres de encabezados
        
        ¿Qué dato regresa y de qué tipo?
        - None
        """
        try:
            from openpyxl.styles import Font, PatternFill, Alignment
            
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
        except ImportError:
            # Si openpyxl no está disponible, agregar headers sin formato
            for col_num, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_num, value=header)
    
    @staticmethod
    def _add_event_rows(
        ws,
        events: List[Dict[str, Any]],
        include_document_details: bool,
        start_row: int = 2
    ) -> None:
        """
        Agrega filas de datos de eventos a la hoja de cálculo.
        
        ¿Qué hace la función?
        Itera sobre los eventos y agrega cada uno como una fila en la hoja,
        formateando fechas y manejando valores nulos.
        
        ¿Qué parámetros recibe y de qué tipo?
        - ws: Worksheet de openpyxl
        - events (List[Dict[str, Any]]): Lista de eventos a agregar
        - include_document_details (bool): Si incluir detalles del documento
        - start_row (int): Fila inicial para datos (default: 2)
        
        ¿Qué dato regresa y de qué tipo?
        - None
        """
        for row_num, event in enumerate(events, start_row):
            col = 1
            
            # ID
            ws.cell(row=row_num, column=col, value=event.get("id") or "")
            col += 1
            
            # Tipo de Evento
            ws.cell(row=row_num, column=col, value=event.get("event_type") or "")
            col += 1
            
            # Descripción
            ws.cell(row=row_num, column=col, value=event.get("description") or "")
            col += 1
            
            # Fecha y Hora
            created_at = event.get("created_at")
            if created_at:
                if isinstance(created_at, datetime):
                    date_str = created_at.strftime("%Y-%m-%d %H:%M:%S")
                else:
                    date_str = str(created_at)
            else:
                date_str = ""
            ws.cell(row=row_num, column=col, value=date_str)
            col += 1
            
            # Detalles del documento (opcional)
            if include_document_details:
                ws.cell(row=row_num, column=col, value=event.get("document_id") or "")
                col += 1
                ws.cell(row=row_num, column=col, value=event.get("document_filename") or "")
                col += 1
                ws.cell(row=row_num, column=col, value=event.get("document_classification") or "")
                col += 1
            
            # ID Usuario
            ws.cell(row=row_num, column=col, value=event.get("user_id") or "")
    
    @staticmethod
    def _auto_adjust_columns(ws, max_width: int = 50) -> None:
        """
        Ajusta automáticamente el ancho de las columnas.
        
        ¿Qué hace la función?
        Calcula el ancho óptimo para cada columna basado en el contenido
        más largo, con un ancho máximo configurable.
        
        ¿Qué parámetros recibe y de qué tipo?
        - ws: Worksheet de openpyxl
        - max_width (int): Ancho máximo permitido para columnas (default: 50)
        
        ¿Qué dato regresa y de qué tipo?
        - None
        """
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, max_width)
            ws.column_dimensions[column].width = adjusted_width

