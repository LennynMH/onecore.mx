"""History use cases."""

import logging
from typing import Optional, Dict, Any
from datetime import datetime
from io import BytesIO

from app.domain.repositories.document_repository import DocumentRepository

logger = logging.getLogger(__name__)


class HistoryUseCases:
    """History use cases."""
    
    def __init__(self, document_repository: DocumentRepository):
        """Initialize use cases with repository."""
        self.document_repository = document_repository
    
    async def get_history(
        self,
        event_type: Optional[str] = None,
        document_id: Optional[int] = None,
        user_id: Optional[int] = None,
        classification: Optional[str] = None,
        date_from: Optional[datetime] = None,
        date_to: Optional[datetime] = None,
        description_search: Optional[str] = None,
        page: int = 1,
        page_size: int = 50
    ) -> Dict[str, Any]:
        """
        Get history with filters and pagination.
        
        Args:
            event_type: Filter by event type
            document_id: Filter by document ID
            user_id: Filter by user ID
            classification: Filter by document classification
            date_from: Filter events from this date
            date_to: Filter events to this date
            description_search: Search in event description
            page: Page number
            page_size: Items per page
            
        Returns:
            Dictionary with events and pagination info
        """
        try:
            result = await self.document_repository.list_events(
                event_type=event_type,
                document_id=document_id,
                user_id=user_id,
                classification=classification,
                date_from=date_from,
                date_to=date_to,
                description_search=description_search,
                page=page,
                page_size=page_size
            )
            
            logger.info(f"History retrieved: {result['total']} total events, page {page}/{result['total_pages']}")
            return result
            
        except Exception as e:
            logger.error(f"Error getting history: {str(e)}")
            raise Exception(f"Failed to get history: {str(e)}")
    
    async def export_to_excel(
        self,
        event_type: Optional[str] = None,
        document_id: Optional[int] = None,
        user_id: Optional[int] = None,
        classification: Optional[str] = None,
        date_from: Optional[datetime] = None,
        date_to: Optional[datetime] = None,
        description_search: Optional[str] = None,
        include_document_details: bool = True
    ) -> BytesIO:
        """
        Export history to Excel.
        
        Args:
            event_type: Filter by event type
            document_id: Filter by document ID
            user_id: Filter by user ID
            classification: Filter by document classification
            date_from: Filter events from this date
            date_to: Filter events to this date
            description_search: Search in event description
            include_document_details: Include document details in export
            
        Returns:
            BytesIO with Excel file content
        """
        try:
            # Get all events (no pagination for export)
            result = await self.document_repository.list_events(
                event_type=event_type,
                document_id=document_id,
                user_id=user_id,
                classification=classification,
                date_from=date_from,
                date_to=date_to,
                description_search=description_search,
                page=1,
                page_size=10000  # Large limit for export
            )
            
            events = result.get("events", [])
            
            # Create Excel file
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment
            except ImportError:
                logger.error("openpyxl not installed. Install it with: pip install openpyxl")
                raise Exception("Excel export requires openpyxl library. Please install it.")
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Historial de Eventos"
            
            # Headers
            headers = ["ID", "Tipo de Evento", "Descripción", "Fecha y Hora"]
            if include_document_details:
                headers.extend(["ID Documento", "Nombre Archivo", "Clasificación"])
            headers.append("ID Usuario")
            
            # Style headers
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Data rows
            for row_num, event in enumerate(events, 2):
                ws.cell(row=row_num, column=1, value=event["id"])
                ws.cell(row=row_num, column=2, value=event["event_type"])
                ws.cell(row=row_num, column=3, value=event["description"])
                ws.cell(row=row_num, column=4, value=event["created_at"].strftime("%Y-%m-%d %H:%M:%S") if event["created_at"] else "")
                
                col = 5
                if include_document_details:
                    ws.cell(row=row_num, column=col, value=event.get("document_id") or "")
                    col += 1
                    ws.cell(row=row_num, column=col, value=event.get("document_filename") or "")
                    col += 1
                    ws.cell(row=row_num, column=col, value=event.get("document_classification") or "")
                    col += 1
                
                ws.cell(row=row_num, column=col, value=event.get("user_id") or "")
            
            # Auto-adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width
            
            # Save to BytesIO
            excel_buffer = BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            logger.info(f"Excel export created: {len(events)} events")
            return excel_buffer
            
        except Exception as e:
            logger.error(f"Error exporting to Excel: {str(e)}")
            raise Exception(f"Failed to export to Excel: {str(e)}")

